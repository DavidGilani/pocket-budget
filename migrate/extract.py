#!/usr/bin/env python3
"""
Pocket Ledger Migration Script
================================
Reads DailyBudgetBackup2.sqlite and finances.xlsx and produces migration-data.json
ready to import into the Pocket Ledger PWA.

Usage:
    python3 extract.py [sqlite_path] [xlsx_path]

Defaults:
    sqlite_path = ../DailyBudgetBackup2.sqlite
    xlsx_path   = ../finances.xlsx

Output:
    migration-data.json (in current directory)
"""

import sys
import json
import sqlite3
import os
from datetime import datetime, timedelta, date

APPLE_EPOCH = datetime(2001, 1, 1)

def apple_ts(v):
    if v is None:
        return None
    try:
        dt = APPLE_EPOCH + timedelta(seconds=v)
        if dt.year > 4001:
            return '4001-01-01'
        return dt.strftime('%Y-%m-%d')
    except:
        return None

CATEGORY_MAP = {
    1: 2,   2: 2,   3: 15,  4: 16,  5: 3,   6: 14,  7: 18,  8: 9,
    9: 21,  10: 6,  11: 19, 12: 7,  13: 24, 14: 3,  15: 10, 16: 9,
    17: 22, 18: 22, 19: 1,  20: 17, 21: 20, 22: 16, 23: 5,  24: 1,
    25: 12, 26: 19, 27: 13, 28: 22, 29: 23, 30: 13, 31: 22, 32: 11,
    33: 8,  34: 4,  35: 22, 36: 7,
}

FG_ACCOUNT_COLS = {
    1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 13,
    10: 7, 11: 9, 12: 8, 13: 10, 14: 11, 15: 12,
}

def migrate_sqlite(conn):
    c = conn.cursor()

    c.execute("""
        SELECT Z_PK, ZDATE, ZAMOUNT, ZCATEGORY, ZNOTE,
               ZISDAILYBUDGET, ZISEXTRAINCOME, ZISWISH, ZISEDITABLE
        FROM ZBOOKING
        WHERE ZISEDITABLE = 1
           OR (ZISEDITABLE IS NULL)
        ORDER BY ZDATE
    """)

    transactions = []
    for row in c.fetchall():
        pk, zdate, amount, cat_legacy, note, is_budget, is_extra, is_wish, is_edit = row
        date_str = apple_ts(zdate)
        if not date_str or amount is None or is_budget == 1:
            continue
        cat_id = CATEGORY_MAP.get(cat_legacy, 22)
        if is_wish:
            txn_type = 'distributed_income' if amount > 0 else 'distributed_expense'
        elif amount > 0:
            txn_type = 'income'
        else:
            txn_type = 'expense'
        transactions.append({
            'id': pk, 'date': date_str, 'amount': round(amount, 2),
            'categoryId': cat_id, 'note': (note or '').strip(), 'type': txn_type,
            'distributionId': None,
            'createdAt': datetime.utcnow().isoformat() + 'Z',
            'updatedAt': datetime.utcnow().isoformat() + 'Z',
            'syncStatus': 'synced',
        })

    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZDESC, ZTIMEOPTION FROM ZFIXEDSPENDING ORDER BY ZSTARTDATE")
    recurring_expenses = []
    now_str = date.today().isoformat()

    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, freq_raw = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or not desc or amount is None:
            continue
        freq_raw = (freq_raw or 'Monthly').strip().lower()
        if 'year' in freq_raw or 'annual' in freq_raw:
            freq = 'yearly'
        elif 'quarter' in freq_raw:
            freq = 'quarterly'
        else:
            freq = 'monthly'
        is_open = (not end or end.startswith('4001') or end > now_str)
        recurring_expenses.append({
            'id': pk, 'description': desc.strip(),
            'amount': round(abs(amount or 0), 2), 'frequency': freq,
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'isShared': False, 'sharePercent': 50,
            'category': None, 'nextReviewDate': None,
            'isActive': bool(is_open),
        })

    SHARED_DESCS = ['mortgage', 'council tax', 'ground rent', 'electricity', 'octopus',
                    'gas', 'water', 'internet', 'netflix', 'disney', 'tv licence',
                    'help to buy', 'htb']
    for r in recurring_expenses:
        if any(s in r['description'].lower() for s in SHARED_DESCS):
            r['isShared'] = True

    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZUSERDESCRIPTION, ZINTERVAL FROM ZREGULARINCOME ORDER BY ZSTARTDATE")
    recurring_income = []
    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, interval = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or amount is None:
            continue
        is_open = (not end or end.startswith('4001') or end > now_str)
        recurring_income.append({
            'id': pk, 'description': (desc or 'Income').strip(),
            'amount': round(abs(amount), 2), 'intervalDays': int(interval or 30),
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'isActive': bool(is_open),
        })

    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZPERCENTAGE FROM ZDB2CYCLESAVING ORDER BY ZSTARTDATE")
    savings_targets = []
    for row in c.fetchall():
        pk, start_ts, end_ts, amount, pct = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start:
            continue
        savings_targets.append({
            'id': pk, 'amount': round(amount or 0, 2), 'percentage': round(pct or 0, 4),
            'startDate': start,
            'endDate': end if end and not end.startswith('4001') else '4001-01-01',
            'description': f'Savings {start[:7]}',
        })

    c.execute("SELECT Z_PK, ZSTARTDATE, ZENDDATE, ZAMOUNT, ZDESC, ZCATEGORY, ZISFINISHED FROM ZWISH ORDER BY ZSTARTDATE")
    distributions = []
    for row in c.fetchall():
        pk, start_ts, end_ts, amount, desc, cat_legacy, finished = row
        start = apple_ts(start_ts)
        end = apple_ts(end_ts)
        if not start or not end or amount is None:
            continue
        distributions.append({
            'id': pk, 'description': (desc or 'Expense').strip(),
            'totalAmount': round(abs(amount), 2), 'startDate': start, 'endDate': end,
            'categoryId': CATEGORY_MAP.get(cat_legacy, 22),
            'isIncome': amount > 0, 'isFinished': bool(finished),
        })

    return transactions, recurring_expenses, recurring_income, savings_targets, distributions


def migrate_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return [], [], []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    account_snapshots = []
    friend_holdings = []
    friend_transactions = []
    snap_id = 1
    ft_id = 1

    if 'Financial goals' in wb.sheetnames:
        ws = wb['Financial goals']
        col_account = {
            0: None, 1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 13,
            8: None, 9: None, 10: 7, 11: 9, 12: 8, 13: 10, 14: 11, 15: 12,
        }
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not row[0]:
                continue
            dt = row[0]
            if isinstance(dt, datetime):
                date_str = dt.strftime('%Y-%m-%d')
            elif isinstance(dt, str):
                date_str = dt[:10]
            else:
                continue
            for col_idx, acc_id in col_account.items():
                if acc_id is None or col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or not isinstance(val, (int, float)):
                    continue
                account_snapshots.append({
                    'id': snap_id, 'accountId': acc_id, 'date': date_str,
                    'balance': round(float(val), 2), 'note': '',
                })
                snap_id += 1

    if 'Bank of Gilulu' in wb.sheetnames:
        ws = wb['Bank of Gilulu']
        friend_holdings.append({'id': 1, 'friendName': 'Dom', 'isActive': True, 'interestRate': 0.04, 'notes': 'Bank of Gilulu'})
        friend_holdings.append({'id': 2, 'friendName': 'Jesse', 'isActive': True, 'interestRate': 0.04, 'notes': 'Bank of Gilulu'})
        for row in ws.iter_rows(min_row=4, values_only=True):
            if len(row) > 3:
                dom_amount, dom_date = row[2], row[3]
                if isinstance(dom_amount, (int, float)) and dom_amount != 0 and isinstance(dom_date, datetime):
                    friend_transactions.append({'id': ft_id, 'holdingId': 1, 'date': dom_date.strftime('%Y-%m-%d'), 'amount': round(float(dom_amount), 2), 'note': ''})
                    ft_id += 1
            if len(row) > 11:
                jesse_amount, jesse_date = row[10], row[11]
                if isinstance(jesse_amount, (int, float)) and jesse_amount != 0 and isinstance(jesse_date, datetime):
                    friend_transactions.append({'id': ft_id, 'holdingId': 2, 'date': jesse_date.strftime('%Y-%m-%d'), 'amount': round(float(jesse_amount), 2), 'note': ''})
                    ft_id += 1

    return account_snapshots, friend_holdings, friend_transactions


def main():
    sqlite_path = sys.argv[1] if len(sys.argv) > 1 else '../DailyBudgetBackup2.sqlite'
    xlsx_path   = sys.argv[2] if len(sys.argv) > 2 else '../finances.xlsx'

    if not os.path.exists(sqlite_path):
        print(f"SQLite file not found: {sqlite_path}")
        sys.exit(1)

    print(f"Reading SQLite: {sqlite_path}")
    conn = sqlite3.connect(sqlite_path)
    transactions, recurring_expenses, recurring_income, savings_targets, distributions = migrate_sqlite(conn)
    conn.close()

    account_snapshots, friend_holdings, friend_transactions = [], [], []
    if os.path.exists(xlsx_path):
        print(f"Reading Excel: {xlsx_path}")
        account_snapshots, friend_holdings, friend_transactions = migrate_xlsx(xlsx_path)
    else:
        print(f"Excel file not found: {xlsx_path} (skipping account snapshots)")

    output = {
        'transactions':      transactions,
        'recurringExpenses': recurring_expenses,
        'recurringIncome':   recurring_income,
        'savingsTargets':    savings_targets,
        'distributions':     distributions,
        'accountSnapshots':  account_snapshots,
        'friendHoldings':    friend_holdings,
        'friendTransactions':friend_transactions,
        'exportedAt': datetime.utcnow().isoformat() + 'Z',
    }

    out_path = 'migration-data.json'
    with open(out_path, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nMigration complete: {out_path}")
    print(f"  {len(transactions)} transactions")
    print(f"  {len(recurring_expenses)} recurring expenses ({sum(1 for r in recurring_expenses if r['isActive'])} active)")
    print(f"  {len(recurring_income)} recurring income records")
    print(f"  {len(savings_targets)} savings targets")
    print(f"  {len(distributions)} distributions")
    print(f"  {len(account_snapshots)} account snapshots")
    print(f"  {len(friend_holdings)} friend holdings, {len(friend_transactions)} transactions")
    print(f"\nImport this file in the app: Settings > Import data")


if __name__ == '__main__':
    main()
